from __future__ import annotations

from pathlib import Path
import json
import re
import unicodedata

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ====================================================
# BEÁLLÍTÁSOK
# ====================================================

INPUT_FOLDER = Path("new_KSH_filled")
OUTPUT_FOLDER = Path("ksh_outputs")

# Excel tartományok
SHEET_NAME = None          # None -> active sheet
HEADER_ROW = 9             # B9:G9 = létszámkategóriák
DATA_START_ROW = 10
DATA_END_ROW = 3199
NAME_COL = 1               # A oszlop = település/kerület
DATA_START_COL = 2         # B
DATA_END_COL = 7           # G

# Kimenetek
LONG_XLSX_NAME = "ksh_long.xlsx"
LONG_CSV_NAME = "ksh_long.csv"
TENSOR_NPY_NAME = "tensor_ksh.npy"
TENSOR_DIMS_JSON_NAME = "tensor_ksh_dimensions.json"

# Létszámkategóriák fix sorrendje
CANONICAL_SIZE_CATS = [
    "1-4 fő",
    "5-9 fő",
    "10-19 fő",
    "20-49 fő",
    "50-249 fő",
    "250 fő felett",
]


# ====================================================
# SZÖVEG / UNICODE NORMALIZÁLÁS
# ====================================================


def normalize_text(value) -> str:
    """Óvatos unicode normalizálás és whitespace tisztítás."""
    if value is None or pd.isna(value):
        return ""

    text = str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text



def fix_common_hungarian_mojibake(text: str) -> str:
    """
    Gyakori hibás latin-1 / cp1250 jellegű karakterek javítása.
    Nem tökéletes minden esetre, de a tipikus õ/û típusú problémákat enyhíti.
    """
    if not text:
        return text

    replacements = {
        "õ": "ő",
        "û": "ű",
        "Õ": "Ő",
        "Û": "Ű",
        "õ": "ő",
        "û": "ű",
        "Õ": "Ő",
        "Û": "Ű",
        "ô": "ő",
        "Ô": "Ő",
        "û": "ű",
        "Û": "Ű",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)

    return text



def normalize_hungarian_text(value) -> str:
    text = normalize_text(value)
    text = fix_common_hungarian_mojibake(text)
    return text



def normalize_size_category(cat) -> str:
    cat = normalize_hungarian_text(cat)
    if not cat:
        return ""

    simplified = re.sub(r"[\s\-]", "", cat.lower())

    for canonical in CANONICAL_SIZE_CATS:
        canonical_simple = re.sub(r"[\s\-]", "", canonical.lower())
        if simplified == canonical_simple:
            return canonical

    # néhány extra biztosítás
    aliases = {
        "14fő": "1-4 fő",
        "59fő": "5-9 fő",
        "1019fő": "10-19 fő",
        "2049fő": "20-49 fő",
        "50249fő": "50-249 fő",
        "250főfelett": "250 fő felett",
        "250felett": "250 fő felett",
    }
    return aliases.get(simplified, cat)


# ====================================================
# FÁJLNÉV FELDOLGOZÁS
# ====================================================


def parse_teaor_from_filename(file_path: Path) -> tuple[str, str]:
    """
    Várt forma pl.:
      C_13_filled.xlsx
      R_91_filled.xlsx
      A_01_filled.xlsx
    Visszaad:
      ("C", "13") vagy ("A", "01")
    """
    match = re.fullmatch(r"([A-Z])_(\d+)_filled", file_path.stem)
    if not match:
        raise ValueError(
            f"A fájlnév nem a várt mintát követi: {file_path.name} "
            f"(elvárt pl. C_13_filled.xlsx)"
        )

    teaor_foag = match.group(1)
    teaor_alag = match.group(2)
    return teaor_foag, teaor_alag


# ====================================================
# EGY EXCEL -> LONG SOROK
# ====================================================


def read_one_ksh_file(file_path: Path) -> tuple[list[dict], list[str], list[str]]:
    """
    Egy feltöltött/feldolgozott KSH Excelből long-format sorokat készít.

    Visszaad:
      rows, territory_order, size_categories_in_sheet
    """
    teaor_foag, teaor_alag = parse_teaor_from_filename(file_path)

    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook[SHEET_NAME] if SHEET_NAME else workbook.active

    size_categories = []
    for col in range(DATA_START_COL, DATA_END_COL + 1):
        raw_cat = worksheet.cell(row=HEADER_ROW, column=col).value
        norm_cat = normalize_size_category(raw_cat)
        size_categories.append(norm_cat)

    if size_categories != CANONICAL_SIZE_CATS:
        raise ValueError(
            f"Hibás vagy váratlan létszámkategóriák a fájlban: {file_path.name}\n"
            f"Talált: {size_categories}\n"
            f"Elvárt: {CANONICAL_SIZE_CATS}"
        )

    rows: list[dict] = []
    territory_order: list[str] = []

    for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
        territory_raw = worksheet.cell(row=row_idx, column=NAME_COL).value
        territory = normalize_hungarian_text(territory_raw)

        if not territory:
            # ha mégis lenne üres sornév, ezt a sort kihagyjuk
            continue

        territory_order.append(territory)

        for offset, col_idx in enumerate(range(DATA_START_COL, DATA_END_COL + 1)):
            size_cat = size_categories[offset]
            value = worksheet.cell(row=row_idx, column=col_idx).value

            if value is None or (isinstance(value, str) and value.strip() == ""):
                value = 0

            value = pd.to_numeric(value, errors="coerce")
            if pd.isna(value):
                value = 0

            rows.append(
                {
                    "Terület": territory,
                    "TEÁOR főág": teaor_foag,
                    "TEÁOR alág kód": teaor_alag,
                    "Létszámkategória": size_cat,
                    "Érték": int(value),
                }
            )

    return rows, territory_order, size_categories


# ====================================================
# ÖSSZES EXCEL -> LONG DATAFRAME
# ====================================================


def natural_teaor_sort_key(file_path: Path):
    teaor_foag, teaor_alag = parse_teaor_from_filename(file_path)
    return teaor_foag, int(teaor_alag)



def build_long_dataframe(input_folder: Path) -> tuple[pd.DataFrame, list[str], list[str], list[str]]:
    files = sorted(input_folder.glob("*_filled.xlsx"), key=natural_teaor_sort_key)

    if not files:
        raise FileNotFoundError(f"Nem találtam *_filled.xlsx fájlokat itt: {input_folder.resolve()}")

    all_rows: list[dict] = []
    territory_order_reference: list[str] | None = None
    teaor_order: list[str] = []

    print(f"Talált fájlok száma: {len(files)}")

    for idx, file_path in enumerate(files, start=1):
        teaor_foag, teaor_alag = parse_teaor_from_filename(file_path)
        teaor_code = f"{teaor_foag}_{teaor_alag}"
        teaor_order.append(teaor_code)

        print(f"[{idx}/{len(files)}] Beolvasás: {file_path.name}")
        rows, territory_order, size_categories = read_one_ksh_file(file_path)

        if territory_order_reference is None:
            territory_order_reference = territory_order
        else:
            if territory_order != territory_order_reference:
                raise ValueError(
                    f"A településsorrend eltér ebben a fájlban: {file_path.name}\n"
                    f"Ez veszélyes a tensor-építésnél, ezért a script megállt."
                )

        all_rows.extend(rows)

    df = pd.DataFrame(all_rows)

    # Stabil sorrendek rögzítése
    territory_order_reference = territory_order_reference or []
    size_order = CANONICAL_SIZE_CATS

    return df, territory_order_reference, size_order, teaor_order


# ====================================================
# LONG DATAFRAME MENTÉSE
# ====================================================


def save_long_outputs(df: pd.DataFrame, output_folder: Path) -> None:
    output_folder.mkdir(parents=True, exist_ok=True)

    csv_path = output_folder / LONG_CSV_NAME

    # CSV: utf-8-sig -> Excel magyar Windows alatt is általában jól kezeli
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"💾 Mentve: {csv_path}")

    # Excel sorlimit kezelése
    max_excel_rows = 1_048_576
    header_rows = 1
    max_data_rows_per_sheet = max_excel_rows - header_rows

    total_rows = len(df)

    if total_rows <= max_data_rows_per_sheet:
        xlsx_path = output_folder / LONG_XLSX_NAME
        df.to_excel(xlsx_path, index=False, engine="openpyxl")
        print(f"💾 Mentve: {xlsx_path}")
    else:
        print(
            f"⚠️ A long DataFrame túl nagy egyetlen Excel munkalaphoz "
            f"({total_rows} sor), ezért több részre bontva mentem."
        )

        base_name = Path(LONG_XLSX_NAME).stem
        suffix = Path(LONG_XLSX_NAME).suffix

        start_row = 0
        part_idx = 1

        while start_row < total_rows:
            end_row = min(start_row + max_data_rows_per_sheet, total_rows)
            chunk = df.iloc[start_row:end_row]

            chunk_path = output_folder / f"{base_name}_part{part_idx}{suffix}"
            chunk.to_excel(chunk_path, index=False, engine="openpyxl")

            print(
                f"💾 Mentve: {chunk_path} "
                f"(sorok: {start_row + 1}-{end_row})"
            )

            start_row = end_row
            part_idx += 1


# ====================================================
# TENZOR ÉPÍTÉS
# ====================================================


def build_tensor_from_long_df(
    df: pd.DataFrame,
    territory_order: list[str],
    size_order: list[str],
    teaor_order: list[str],
    output_folder: Path,
) -> np.ndarray:
    """
    Tensor shape:
      (terület, létszámkategória, teáor)
    ahol a teáor dimenzió elemei pl. A_01, C_13, R_91.
    """
    df = df.copy()
    df["TEÁOR_kulcs"] = df["TEÁOR főág"] + "_" + df["TEÁOR alág kód"].astype(str)

    # Pivot: sorok = területek, oszlopok = (létszámbin, teáor)
    pivot = df.pivot_table(
        index="Terület",
        columns=["Létszámkategória", "TEÁOR_kulcs"],
        values="Érték",
        aggfunc="sum",
        fill_value=0,
    )

    target_columns = pd.MultiIndex.from_product(
        [size_order, teaor_order],
        names=["Létszámkategória", "TEÁOR_kulcs"],
    )

    pivot = pivot.reindex(index=territory_order, columns=target_columns, fill_value=0)

    tensor = pivot.to_numpy(dtype=np.int32).reshape(
        len(territory_order),
        len(size_order),
        len(teaor_order),
    )

    npy_path = output_folder / TENSOR_NPY_NAME
    dims_path = output_folder / TENSOR_DIMS_JSON_NAME

    np.save(npy_path, tensor)

    dims = {
        "teruletek": territory_order,
        "letszamkategoriak": size_order,
        "teaor_kulcsok": teaor_order,
        "teaor_feloldas": [
            {
                "teaor_kulcs": code,
                "TEÁOR főág": code.split("_")[0],
                "TEÁOR alág kód": code.split("_")[1],
            }
            for code in teaor_order
        ],
        "shape": {
            "teruletek": len(territory_order),
            "letszamkategoriak": len(size_order),
            "teaorok": len(teaor_order),
        },
    }

    with open(dims_path, "w", encoding="utf-8") as f:
        json.dump(dims, f, ensure_ascii=False, indent=2)

    print(f"✅ Tenzor shape: {tensor.shape}")
    print(f"💾 Mentve: {npy_path}")
    print(f"💾 Mentve: {dims_path}")

    return tensor


# ====================================================
# ELLENŐRZÉS / FUTTATÁS
# ====================================================


def main():
    print("📥 KSH fájlok beolvasása indul...")

    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    df, territory_order, size_order, teaor_order = build_long_dataframe(INPUT_FOLDER)

    print("\n📊 Long DataFrame elkészült")
    print(f"Sorok száma: {len(df):,}")
    print(f"Területek száma: {len(territory_order)}")
    print(f"Létszámkategóriák száma: {len(size_order)}")
    print(f"TEÁOR-ok száma: {len(teaor_order)}")

    save_long_outputs(df, OUTPUT_FOLDER)
    build_tensor_from_long_df(df, territory_order, size_order, teaor_order, OUTPUT_FOLDER)

    print("\n🎉 Kész: összesített tábla + tensor legenerálva.")


if __name__ == "__main__":
    main()
