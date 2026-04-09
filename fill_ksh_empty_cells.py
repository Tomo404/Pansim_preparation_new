from pathlib import Path
from openpyxl import load_workbook


INPUT_FOLDER = Path("new_KSH")
OUTPUT_FOLDER = Path("new_KSH_filled")

DATA_START_ROW = 10
DATA_END_ROW = 3199

DATA_START_COL = 2   # B oszlop
DATA_END_COL = 7     # G oszlop


def is_empty_cell(value):
    """
    Eldönti, hogy egy cellát üresnek tekintünk-e.
    Üresnek számít:
    - None
    - üres string
    - csak whitespace-et tartalmazó string
    """
    if value is None:
        return True

    if isinstance(value, str) and value.strip() == "":
        return True

    return False


def fill_empty_cells_in_workbook(input_file: Path, output_file: Path):
    """
    Beolvassa az input Excel fájlt, a B10:G3199 tartományban
    az üres cellákat 0-ra cseréli, majd elmenti az output fájlba.
    """
    workbook = load_workbook(input_file)
    worksheet = workbook.active

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        for col in range(DATA_START_COL, DATA_END_COL + 1):
            cell = worksheet.cell(row=row, column=col)

            if is_empty_cell(cell.value):
                cell.value = 0

    workbook.save(output_file)


def main():
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    excel_files = sorted(INPUT_FOLDER.glob("*.xlsx"))

    if not excel_files:
        print(f"Nincs .xlsx fájl az input mappában: {INPUT_FOLDER.resolve()}")
        return

    print(f"Talált fájlok száma: {len(excel_files)}")

    for input_file in excel_files:
        output_name = f"{input_file.stem}_filled{input_file.suffix}"
        output_file = OUTPUT_FOLDER / output_name

        print(f"Feldolgozás: {input_file.name} -> {output_name}")
        fill_empty_cells_in_workbook(input_file, output_file)

    print("Kész. Minden fájl feldolgozva.")


if __name__ == "__main__":
    main()